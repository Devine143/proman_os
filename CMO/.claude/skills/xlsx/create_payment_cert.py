from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
sheet = wb.active
sheet.title = "Payment Certificate"

# Set column widths
sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['B'].width = 35
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 18

# Styles
header_font = Font(name='Arial', size=14, bold=True)
subheader_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
blue_font = Font(name='Arial', size=10, color='0000FF')  # Blue for inputs
amount_font = Font(name='Arial', size=10, bold=True)
total_font = Font(name='Arial', size=11, bold=True)

header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
section_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
total_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
wrap_align = Alignment(wrap_text=True, vertical='top')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# HEADER
sheet.merge_cells('A1:D1')
sheet['A1'] = 'PAYMENT CERTIFICATE'
sheet['A1'].font = Font(name='Arial', size=16, bold=True)
sheet['A1'].alignment = center_align
sheet['A1'].fill = header_fill

sheet.merge_cells('A2:D2')
sheet['A2'] = 'JBCC Principal Building Agreement'
sheet['A2'].font = subheader_font
sheet['A2'].alignment = center_align
sheet['A2'].fill = header_fill

# Certificate Type
sheet.merge_cells('A3:D3')
sheet['A3'] = 'INTERIM PAYMENT CERTIFICATE'
sheet['A3'].font = Font(name='Arial', size=12, bold=True, color='FF0000')
sheet['A3'].alignment = center_align

# PROJECT INFORMATION
row = 5
sheet[f'A{row}'] = 'PROJECT INFORMATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Project Name:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[Enter Project Name]'
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill

row += 1
sheet[f'A{row}'] = 'Project Address:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[Enter Project Address]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Employer:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[Enter Employer Name]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Contractor:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[Enter Contractor Name]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Contract Date:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[Enter Contract Date]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Contract Price (Excl VAT):'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = yellow_fill

# CERTIFICATE DETAILS
row += 2
sheet[f'A{row}'] = 'CERTIFICATE DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Certificate Number:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[e.g., PC-001]'
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill

row += 1
sheet[f'A{row}'] = 'Valuation Period:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[e.g., Month ending 31 Jan 2025]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Date of Certificate:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = datetime.now().strftime('%d %B %Y')
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill

row += 1
sheet[f'A{row}'] = 'Payment Due Date:'
sheet[f'A{row}'].font = normal_font
sheet[f'B{row}'] = '[21 days from certificate date]'
sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True, color='FF0000')
sheet[f'B{row}'].fill = yellow_fill

# VALUATION SECTION
val_start_row = row + 2
row = val_start_row
sheet[f'A{row}'] = 'VALUATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

# Table headers
row += 1
headers_row = row
sheet[f'A{row}'] = 'Description'
sheet[f'B{row}'] = 'Contract Value'
sheet[f'C{row}'] = 'Value This Period'
sheet[f'D{row}'] = 'Cumulative Value'
for col in ['A', 'B', 'C', 'D']:
    cell = sheet[f'{col}{row}']
    cell.font = subheader_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Line items
row += 1
line1_row = row
sheet[f'A{row}'] = '1. Original Contract Works'
sheet[f'B{row}'] = '=B12'
sheet[f'C{row}'] = 0
sheet[f'D{row}'] = 0
sheet[f'C{row}'].font = blue_font
sheet[f'C{row}'].fill = yellow_fill
sheet[f'D{row}'].font = blue_font
sheet[f'D{row}'].fill = yellow_fill

row += 1
sheet[f'A{row}'] = '2. Approved Variations'
sheet[f'B{row}'] = 0
sheet[f'C{row}'] = 0
sheet[f'D{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'C{row}'].font = blue_font
sheet[f'D{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = '3. Provisional Sums Expended'
sheet[f'B{row}'] = 0
sheet[f'C{row}'] = 0
sheet[f'D{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'C{row}'].font = blue_font
sheet[f'D{row}'].font = blue_font

row += 1
subtotal_row = row
sheet[f'A{row}'] = 'SUBTOTAL - Work Value (Excl VAT)'
sheet[f'A{row}'].font = amount_font
sheet[f'B{row}'] = f'=SUM(B{line1_row}:B{row-1})'
sheet[f'C{row}'] = f'=SUM(C{line1_row}:C{row-1})'
sheet[f'D{row}'] = f'=SUM(D{line1_row}:D{row-1})'
sheet[f'A{row}'].fill = section_fill

row += 1
sheet[f'A{row}'] = '4. Materials on Site (if applicable)'
sheet[f'B{row}'] = 0
sheet[f'C{row}'] = 0
sheet[f'D{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'C{row}'].font = blue_font
sheet[f'D{row}'].font = blue_font

row += 1
gross_val_row = row
sheet[f'A{row}'] = 'GROSS VALUE (Excl VAT)'
sheet[f'A{row}'].font = amount_font
sheet[f'B{row}'] = f'=B{subtotal_row}+B{row-1}'
sheet[f'C{row}'] = f'=C{subtotal_row}+C{row-1}'
sheet[f'D{row}'] = f'=D{subtotal_row}+D{row-1}'
sheet[f'A{row}'].fill = section_fill

# DEDUCTIONS
row += 2
deduct_start_row = row
sheet[f'A{row}'] = 'DEDUCTIONS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
retention_rate_row = row
sheet[f'A{row}'] = 'Retention Rate:'
sheet[f'A{row}'].font = normal_font
sheet[f'C{row}'] = 0.05
sheet[f'C{row}'].font = blue_font
sheet[f'C{row}'].number_format = '0.0%'
sheet[f'C{row}'].fill = yellow_fill
sheet[f'D{row}'] = '(Typically 5%)'
sheet[f'D{row}'].font = Font(name='Arial', size=9, italic=True)

row += 1
retention_row = row
sheet[f'A{row}'] = 'Less: Retention'
sheet[f'D{row}'] = f'=D{gross_val_row}*$C${retention_rate_row}'
sheet[f'D{row}'].font = Font(name='Arial', size=10, color='FF0000')

row += 1
prev_cert_row = row
sheet[f'A{row}'] = 'Less: Previous Certificates'
sheet[f'D{row}'] = 0
sheet[f'D{row}'].font = blue_font
sheet[f'D{row}'].fill = yellow_fill

row += 1
net_excl_vat_row = row
sheet[f'A{row}'] = 'NET AMOUNT DUE (Excl VAT)'
sheet[f'A{row}'].font = total_font
sheet[f'D{row}'] = f'=D{gross_val_row}-D{retention_row}-D{prev_cert_row}'
sheet[f'D{row}'].font = total_font
sheet[f'A{row}'].fill = total_fill
sheet[f'D{row}'].fill = total_fill

# VAT
row += 2
vat_rate_row = row
sheet[f'A{row}'] = 'VAT Rate:'
sheet[f'A{row}'].font = normal_font
sheet[f'C{row}'] = 0.15
sheet[f'C{row}'].font = blue_font
sheet[f'C{row}'].number_format = '0.0%'
sheet[f'C{row}'].fill = yellow_fill
sheet[f'D{row}'] = '(Currently 15% in SA)'
sheet[f'D{row}'].font = Font(name='Arial', size=9, italic=True)

row += 1
vat_row = row
sheet[f'A{row}'] = 'Add: VAT'
sheet[f'D{row}'] = f'=D{net_excl_vat_row}*$C${vat_rate_row}'

row += 2
total_row = row
sheet[f'A{row}'] = 'TOTAL AMOUNT DUE (Incl VAT)'
sheet[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
sheet[f'D{row}'] = f'=D{net_excl_vat_row}+D{vat_row}'
sheet[f'D{row}'].font = Font(name='Arial', size=12, bold=True)
sheet[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sheet[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
sheet[f'D{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sheet[f'D{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')

# Apply borders and number formatting to valuation table
for r in range(headers_row, gross_val_row + 1):
    for col in ['A', 'B', 'C', 'D']:
        cell = sheet[f'{col}{r}']
        cell.border = thin_border
        if col in ['B', 'C', 'D'] and r > headers_row:
            cell.number_format = 'R #,##0.00'
            
# Apply formatting to deductions
for r in range(retention_row, total_row + 1):
    for col in ['A', 'D']:
        cell = sheet[f'{col}{r}']
        if col == 'D' and r != total_row:
            cell.number_format = 'R #,##0.00'

# PAYMENT INSTRUCTIONS
row += 3
sheet[f'A{row}'] = 'PAYMENT DUE DATE & LEGAL NOTICE'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row+2}')
legal_text = (
    "Payment of the certified amount is DUE WITHIN 21 DAYS from the date of this certificate "
    "in accordance with Clause 6.25 of the JBCC Principal Building Agreement.\n\n"
    "This Payment Certificate constitutes a LIQUID ACKNOWLEDGMENT OF DEBT and is BINDING AND "
    "EQUIVALENT TO CASH (Joob Joob Investments v Stocks Mavundla Zek JV, 2009 SCA). "
    "Failure to pay may result in contractor suspension of works and/or legal proceedings."
)
sheet[f'A{row}'] = legal_text
sheet[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color='FF0000')
sheet[f'A{row}'].alignment = wrap_align
sheet[f'A{row}'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
sheet.row_dimensions[row].height = 60

# BANKING DETAILS
row += 3
sheet[f'A{row}'] = 'CONTRACTOR BANKING DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Bank Name:'
sheet[f'B{row}'] = '[Enter Bank Name]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Account Name:'
sheet[f'B{row}'] = '[Enter Account Name]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Account Number:'
sheet[f'B{row}'] = '[Enter Account Number]'
sheet[f'B{row}'].font = blue_font

row += 1
sheet[f'A{row}'] = 'Branch Code:'
sheet[f'B{row}'] = '[Enter Branch Code]'
sheet[f'B{row}'].font = blue_font

# CERTIFICATION
row += 3
sheet[f'A{row}'] = 'PRINCIPAL AGENT CERTIFICATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row+1}')
cert_text = (
    "I hereby certify that I have inspected the works and reviewed the Quantity Surveyor's "
    "valuation, and that the amount stated above represents a fair and accurate valuation "
    "of works executed and materials on site as at the valuation date, in accordance with "
    "the JBCC Principal Building Agreement."
)
sheet[f'A{row}'] = cert_text
sheet[f'A{row}'].font = normal_font
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 45

row += 2
sheet[f'A{row}'] = 'Principal Agent Name:'
sheet[f'B{row}'] = '[Enter PA Name]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Company:'
sheet[f'B{row}'] = 'Proman Project Management'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Signature:'
sheet[f'B{row}'] = '________________________'
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Date:'
sheet[f'B{row}'] = datetime.now().strftime('%d %B %Y')
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:D{row}')

# FOOTER NOTE
row += 3
sheet.merge_cells(f'A{row}:D{row+1}')
footer_text = (
    "NOTES:\n"
    "1. This certificate is issued under Clause 6.25 of the JBCC Principal Building Agreement\n"
    "2. Retention to be released: 50% at Practical Completion, 50% at Final Completion (per contract terms)\n"
    "3. Employer must pay within 21 days - cannot withhold payment due to defects (Group Five v IDT, 2025)\n"
    "4. Any disputes regarding this certificate must be raised within 10 working days (JBCC adjudication rules)"
)
sheet[f'A{row}'] = footer_text
sheet[f'A{row}'].font = Font(name='Arial', size=8, italic=True)
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 55

# Save
wb.save('/Users/donovan/Archive/AGENT/templates/JBCC_Payment_Certificate_Template.xlsx')
print("Payment Certificate template created successfully")
