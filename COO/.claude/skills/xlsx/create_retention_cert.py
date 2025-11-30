from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = Workbook()
sheet = wb.active
sheet.title = "Retention Release"

sheet.column_dimensions['A'].width = 32
sheet.column_dimensions['B'].width = 35
sheet.column_dimensions['C'].width = 18

header_font = Font(name='Arial', size=14, bold=True)
subheader_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
blue_font = Font(name='Arial', size=10, color='0000FF')
total_font = Font(name='Arial', size=11, bold=True)

header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
section_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(wrap_text=True, vertical='top')

# HEADER
sheet.merge_cells('A1:C1')
sheet['A1'] = 'RETENTION RELEASE CERTIFICATE'
sheet['A1'].font = Font(name='Arial', size=16, bold=True)
sheet['A1'].alignment = center_align
sheet['A1'].fill = header_fill

sheet.merge_cells('A2:C2')
sheet['A2'] = 'JBCC Principal Building Agreement'
sheet['A2'].font = subheader_font
sheet['A2'].alignment = center_align
sheet['A2'].fill = header_fill

# Release Type Selection
row = 4
sheet.merge_cells(f'A{row}:C{row}')
sheet[f'A{row}'] = 'RETENTION RELEASE TYPE (Select one):'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = yellow_fill

row += 1
sheet[f'A{row}'] = '☐ Partial Release at Practical Completion (50%)'
sheet[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
sheet.merge_cells(f'A{row}:C{row}')

row += 1
sheet[f'A{row}'] = '☐ Final Release at Final Completion (Remaining 50%)'
sheet[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
sheet.merge_cells(f'A{row}:C{row}')

# PROJECT INFO
row += 2
sheet[f'A{row}'] = 'PROJECT INFORMATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

for label, value in [
    ('Project Name:', '[Enter Project Name]'),
    ('Project Address:', '[Enter Project Address]'),
    ('Employer:', '[Enter Employer Name]'),
    ('Contractor:', '[Enter Contractor Name]'),
    ('Contract Date:', '[Enter Contract Date]')
]:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'A{row}'].font = normal_font
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font
    sheet.merge_cells(f'B{row}:C{row}')

# CERTIFICATE DETAILS
row += 2
sheet[f'A{row}'] = 'CERTIFICATE DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

row += 1
sheet[f'A{row}'] = 'Retention Release Certificate Number:'
sheet[f'B{row}'] = 'RR-001'
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill
sheet.merge_cells(f'B{row}:C{row}')

row += 1
sheet[f'A{row}'] = 'Date of This Certificate:'
sheet[f'B{row}'] = datetime.now().strftime('%d %B %Y')
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill
sheet.merge_cells(f'B{row}:C{row}')

row += 1
sheet[f'A{row}'] = 'Practical Completion Certificate No.:'
sheet[f'B{row}'] = '[Enter PC Cert No.]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:C{row}')

row += 1
sheet[f'A{row}'] = 'Practical Completion Date:'
sheet[f'B{row}'] = '[Enter PC Date]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:C{row}')

row += 1
sheet[f'A{row}'] = 'Final Completion Certificate No. (if applicable):'
sheet[f'B{row}'] = '[Enter FC Cert No. or N/A]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:C{row}')

row += 1
sheet[f'A{row}'] = 'Final Completion Date (if applicable):'
sheet[f'B{row}'] = '[Enter FC Date or N/A]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:C{row}')

# RETENTION CALCULATION
row += 2
sheet[f'A{row}'] = 'RETENTION CALCULATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

row += 1
contract_val_row = row
sheet[f'A{row}'] = 'Final Contract Value (Excl VAT):'
sheet[f'B{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = yellow_fill

row += 1
retention_rate_row = row
sheet[f'A{row}'] = 'Retention Rate:'
sheet[f'B{row}'] = 0.05
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].number_format = '0.0%'
sheet[f'B{row}'].fill = yellow_fill
sheet[f'C{row}'] = '(Typically 5%)'
sheet[f'C{row}'].font = Font(name='Arial', size=9, italic=True)

row += 1
total_retention_row = row
sheet[f'A{row}'] = 'Total Retention (Contract × Rate):'
sheet[f'B{row}'] = f'=B{contract_val_row}*B{retention_rate_row}'
sheet[f'B{row}'].font = total_font
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = section_fill

row += 2
sheet[f'A{row}'] = 'RETENTION STATUS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

row += 1
prev_released_row = row
sheet[f'A{row}'] = 'Previously Released Retention:'
sheet[f'B{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = yellow_fill

row += 1
this_release_row = row
sheet[f'A{row}'] = 'Retention Released This Certificate:'
sheet[f'B{row}'] = 0
sheet[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='008000')
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = yellow_fill

row += 1
remaining_row = row
sheet[f'A{row}'] = 'Retention Remaining:'
sheet[f'B{row}'] = f'=B{total_retention_row}-B{prev_released_row}-B{this_release_row}'
sheet[f'B{row}'].font = Font(name='Arial', size=10, color='FF0000')
sheet[f'B{row}'].number_format = 'R #,##0.00'

# PAYMENT CALCULATION
row += 2
sheet[f'A{row}'] = 'PAYMENT DUE'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

row += 1
net_excl_vat_row = row
sheet[f'A{row}'] = 'Retention Release Amount (Excl VAT):'
sheet[f'B{row}'] = f'=B{this_release_row}'
sheet[f'B{row}'].font = total_font
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = green_fill

row += 1
vat_rate_row = row
sheet[f'A{row}'] = 'VAT Rate:'
sheet[f'B{row}'] = 0.15
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].number_format = '0.0%'
sheet[f'B{row}'].fill = yellow_fill

row += 1
vat_row = row
sheet[f'A{row}'] = 'Add: VAT:'
sheet[f'B{row}'] = f'=B{net_excl_vat_row}*B{vat_rate_row}'
sheet[f'B{row}'].number_format = 'R #,##0.00'

row += 2
total_row = row
sheet[f'A{row}'] = 'TOTAL RETENTION RELEASE (Incl VAT):'
sheet[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
sheet[f'B{row}'] = f'=B{net_excl_vat_row}+B{vat_row}'
sheet[f'B{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sheet[f'B{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# PAYMENT DETAILS
row += 2
sheet[f'A{row}'] = 'PAYMENT DUE DATE:'
sheet[f'A{row}'].font = Font(name='Arial', size=11, bold=True, color='FF0000')
sheet[f'B{row}'] = '[21 days from this certificate]'
sheet[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='FF0000')
sheet[f'B{row}'].fill = yellow_fill
sheet.merge_cells(f'B{row}:C{row}')

# LEGAL NOTICE
row += 2
sheet[f'A{row}'] = 'LEGAL NOTICE'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

row += 1
sheet.merge_cells(f'A{row}:C{row+2}')
legal_text = (
    "This Retention Release Certificate constitutes a PAYMENT CERTIFICATE under the JBCC "
    "Principal Building Agreement and is BINDING AND EQUIVALENT TO CASH (Joob Joob Investments "
    "v Stocks Mavundla, 2009 SCA).\n\n"
    "Payment is DUE WITHIN 21 DAYS from the date of this certificate. Failure to pay may "
    "result in contractor suspension of works and/or legal proceedings.\n\n"
    "AT PRACTICAL COMPLETION: Typically 50% of retention is released (verify contract terms).\n"
    "AT FINAL COMPLETION: Remaining retention is released after all defects rectified."
)
sheet[f'A{row}'] = legal_text
sheet[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color='FF0000')
sheet[f'A{row}'].alignment = wrap_align
sheet[f'A{row}'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
sheet.row_dimensions[row].height = 90

# CONDITIONS FOR RELEASE
row += 3
sheet[f'A{row}'] = 'CONDITIONS VERIFIED FOR RELEASE'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

conditions = [
    ('☐', 'Practical Completion Certificate issued'),
    ('☐', 'Final Completion Certificate issued (for final 50% release)'),
    ('☐', 'All defects on snag list rectified (for final release)'),
    ('☐', 'As-built drawings submitted'),
    ('☐', 'Operation & Maintenance manuals submitted'),
    ('☐', 'Guarantees and warranties submitted'),
    ('☐', 'No outstanding claims against contractor')
]

for checkbox, condition in conditions:
    row += 1
    sheet[f'A{row}'] = checkbox
    sheet[f'B{row}'] = condition
    sheet[f'B{row}'].font = normal_font
    sheet.merge_cells(f'B{row}:C{row}')

# BANKING DETAILS
row += 2
sheet[f'A{row}'] = 'CONTRACTOR BANKING DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

for label in ['Bank Name:', 'Account Name:', 'Account Number:', 'Branch Code:']:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'B{row}'] = f'[Enter {label.replace(":", "")}]'
    sheet[f'B{row}'].font = blue_font
    sheet.merge_cells(f'B{row}:C{row}')

# CERTIFICATION
row += 2
sheet[f'A{row}'] = 'PRINCIPAL AGENT CERTIFICATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:C{row}')

row += 1
sheet.merge_cells(f'A{row}:C{row+2}')
cert_text = (
    "I hereby certify that:\n\n"
    "1. The applicable Completion Certificate (Practical or Final) has been issued\n"
    "2. The conditions for retention release have been verified and satisfied\n"
    "3. The amount stated above is due and payable to the Contractor\n"
    "4. This certificate is issued in accordance with the JBCC Principal Building Agreement"
)
sheet[f'A{row}'] = cert_text
sheet[f'A{row}'].font = normal_font
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 70

row += 3
for label, value in [
    ('Principal Agent Name:', '[Enter PA Name]'),
    ('Company:', 'Proman Project Management'),
    ('Signature:', '________________________'),
    ('Date:', datetime.now().strftime('%d %B %Y'))
]:
    sheet[f'A{row}'] = label
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font if '[' in str(value) or 'Proman' in str(value) or datetime.now().strftime('%Y') in str(value) else normal_font
    sheet.merge_cells(f'B{row}:C{row}')
    row += 1

# FOOTER
row += 1
sheet.merge_cells(f'A{row}:C{row+1}')
footer = (
    "NOTES:\n"
    "1. This certificate creates immediate payment obligation within 21 days\n"
    "2. Retention percentages and release timing per contract terms (verify JBCC clause)\n"
    "3. Final retention not released until all defects rectified and documentation submitted\n"
    "4. Employer may withhold retention if contractor fails to rectify defects during DLP"
)
sheet[f'A{row}'] = footer
sheet[f'A{row}'].font = Font(name='Arial', size=8, italic=True)
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 50

wb.save('/Users/donovan/Archive/AGENT/templates/JBCC_Retention_Release_Certificate_Template.xlsx')
print("Retention Release Certificate template created")
