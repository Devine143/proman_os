from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

wb = Workbook()
sheet = wb.active
sheet.title = "Practical Completion Cert"

sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15

header_font = Font(name='Arial', size=14, bold=True)
subheader_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
blue_font = Font(name='Arial', size=10, color='0000FF')

header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
section_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# HEADER
sheet.merge_cells('A1:D1')
sheet['A1'] = 'PRACTICAL COMPLETION CERTIFICATE'
sheet['A1'].font = Font(name='Arial', size=16, bold=True)
sheet['A1'].alignment = center_align
sheet['A1'].fill = header_fill

sheet.merge_cells('A2:D2')
sheet['A2'] = 'JBCC Principal Building Agreement - Clause 6.27'
sheet['A2'].font = subheader_font
sheet['A2'].alignment = center_align
sheet['A2'].fill = header_fill

sheet.merge_cells('A3:D3')
sheet['A3'] = 'Certification of Substantial Completion'
sheet['A3'].font = Font(name='Arial', size=11, bold=True, color='008000')
sheet['A3'].alignment = center_align

# PROJECT INFO
row = 5
sheet[f'A{row}'] = 'PROJECT INFORMATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

for label, value in [
    ('Project Name:', '[Enter Project Name]'),
    ('Project Address:', '[Enter Project Address]'),
    ('Employer:', '[Enter Employer Name]'),
    ('Contractor:', '[Enter Contractor Name]'),
    ('Contract Date:', '[Enter Contract Date]'),
    ('Contract Sum (Excl VAT):', 0)
]:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'A{row}'].font = normal_font
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font
    if 'Sum' in label:
        sheet[f'B{row}'].number_format = 'R #,##0.00'
        sheet[f'B{row}'].fill = yellow_fill
    sheet.merge_cells(f'B{row}:D{row}')

# COMPLETION DETAILS
row += 2
sheet[f'A{row}'] = 'COMPLETION DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Certificate Number:'
sheet[f'B{row}'] = 'PC-001'
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Date of Practical Completion:'
sheet[f'B{row}'] = datetime.now().strftime('%d %B %Y')
sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True, color='008000')
sheet[f'B{row}'].fill = yellow_fill
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Original Completion Date per Contract:'
sheet[f'B{row}'] = '[Enter Original Date]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Adjusted Completion Date (with EOTs):'
sheet[f'B{row}'] = '[Enter Adjusted Date]'
sheet[f'B{row}'].font = blue_font
sheet.merge_cells(f'B{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Defects Liability Period:'
sheet[f'B{row}'] = '90 days'
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].fill = yellow_fill
sheet[f'C{row}'] = '(Verify contract terms)'
sheet[f'C{row}'].font = Font(name='Arial', size=9, italic=True)

row += 1
dlp_end_row = row
sheet[f'A{row}'] = 'Defects Liability Period Ends:'
sheet[f'B{row}'] = '[PC Date + DLP days]'
sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True, color='FF0000')
sheet[f'B{row}'].fill = yellow_fill
sheet.merge_cells(f'B{row}:D{row}')

# CERTIFICATION STATEMENT
row += 2
sheet[f'A{row}'] = 'PRINCIPAL AGENT CERTIFICATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row+3}')
cert_statement = (
    "I hereby certify that:\n\n"
    "1. The works have reached PRACTICAL COMPLETION as defined in the JBCC Principal Building Agreement\n"
    "2. The building is SUBSTANTIALLY COMPLETE and fit for occupation/use for its intended purpose\n"
    "3. Minor outstanding works and defects are noted on the attached Snag List\n"
    "4. The works comply with the Contract Documents in all material respects\n\n"
    "This certificate is issued in accordance with Clause 6.27 of the JBCC Principal Building Agreement."
)
sheet[f'A{row}'] = cert_statement
sheet[f'A{row}'].font = normal_font
sheet[f'A{row}'].alignment = wrap_align
sheet[f'A{row}'].fill = green_fill
sheet.row_dimensions[row].height = 100

# SNAG LIST
row += 4
sheet[f'A{row}'] = 'SNAG LIST - Outstanding Items'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row}')
snag_note = (
    "Note: Practical Completion can be issued with minor defects (Group Five v IDT, 2025). "
    "The building must be substantially complete and usable. Major defects prevent PC certification."
)
sheet[f'A{row}'] = snag_note
sheet[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color='0000FF')
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 30

# Snag list table headers
row += 1
snag_header_row = row
for col, text in [('A', 'Item #'), ('B', 'Description'), ('C', 'Location'), ('D', 'Target Date')]:
    sheet[f'{col}{row}'] = text
    sheet[f'{col}{row}'].font = subheader_font
    sheet[f'{col}{row}'].fill = header_fill
    sheet[f'{col}{row}'].alignment = center_align
    sheet[f'{col}{row}'].border = thin_border

# Sample snag items
snag_items = [
    ('1', 'Touch up paint - main entrance', 'Ground Floor Foyer', ''),
    ('2', 'Replace cracked floor tile', 'Office 201', ''),
    ('3', 'Adjust door closer', 'Conference Room', ''),
    ('4', 'Seal gap at window frame', 'Stairwell Level 2', ''),
    ('5', '', '', '')
]

for item_num, desc, loc, target in snag_items:
    row += 1
    sheet[f'A{row}'] = item_num
    sheet[f'B{row}'] = desc
    sheet[f'C{row}'] = loc
    sheet[f'D{row}'] = target
    for col in ['A', 'B', 'C', 'D']:
        sheet[f'{col}{row}'].border = thin_border
        if desc:
            sheet[f'{col}{row}'].font = Font(name='Arial', size=9)
        else:
            sheet[f'{col}{row}'].fill = yellow_fill

snag_end_row = row

row += 1
sheet.merge_cells(f'A{row}:D{row}')
sheet[f'A{row}'] = '[Add additional rows as needed for complete snag list]'
sheet[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color='0000FF')

# LEGAL NOTICE
row += 2
sheet[f'A{row}'] = 'LEGAL CONSEQUENCES OF PRACTICAL COMPLETION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row+2}')
legal_text = (
    "Issuance of Practical Completion Certificate triggers:\n\n"
    "1. START of Defects Liability Period (DLP) - contractor must rectify snag list items during DLP\n"
    "2. RETENTION RELEASE - Typically 50% of retention is released (per contract terms)\n"
    "3. EMPLOYER OCCUPATION - Employer may occupy/use the building\n"
    "4. INSURANCE RESPONSIBILITY - May transfer from contractor to employer (verify insurance terms)\n"
    "5. PAYMENT NOT SUSPENDED - Payment certificates remain enforceable despite snag list "
    "(Group Five v IDT, 2025)\n\n"
    "Defects on snag list must be rectified during DLP. If not rectified, Employer may engage "
    "another contractor at original contractor's cost (Clause 6.27)."
)
sheet[f'A{row}'] = legal_text
sheet[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color='FF0000')
sheet[f'A{row}'].alignment = wrap_align
sheet[f'A{row}'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
sheet.row_dimensions[row].height = 110

# RETENTION RELEASE
row += 3
sheet[f'A{row}'] = 'RETENTION RELEASE (If applicable per contract)'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'Total Retention Held:'
sheet[f'B{row}'] = 0
sheet[f'B{row}'].font = blue_font
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].fill = yellow_fill

row += 1
sheet[f'A{row}'] = 'Retention Released at PC (50%):'
sheet[f'B{row}'] = '=B' + str(row-1) + '*0.5'
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].font = Font(name='Arial', size=10, color='008000')

row += 1
sheet[f'A{row}'] = 'Retention Remaining (50%):'
sheet[f'B{row}'] = '=B' + str(row-2) + '*0.5'
sheet[f'B{row}'].number_format = 'R #,##0.00'
sheet[f'B{row}'].font = Font(name='Arial', size=10, color='FF0000')

row += 1
sheet[f'A{row}'] = 'Remaining retention released at Final Completion'
sheet[f'A{row}'].font = Font(name='Arial', size=9, italic=True)
sheet.merge_cells(f'A{row}:D{row}')

# SIGNATURES
row += 2
sheet[f'A{row}'] = 'SIGNATURES'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet[f'A{row}'] = 'PRINCIPAL AGENT'
sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
sheet.merge_cells(f'A{row}:D{row}')

for label, value in [
    ('Name:', '[Enter PA Name]'),
    ('Company:', 'Proman Project Management'),
    ('Signature:', '________________________'),
    ('Date:', datetime.now().strftime('%d %B %Y'))
]:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font if '[' in str(value) or 'Proman' in str(value) or datetime.now().strftime('%Y') in str(value) else normal_font
    sheet.merge_cells(f'B{row}:D{row}')

row += 2
sheet[f'A{row}'] = 'CONTRACTOR ACKNOWLEDGMENT'
sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
sheet.merge_cells(f'A{row}:D{row}')

for label in ['Name:', 'Signature:', 'Date:']:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'B{row}'] = '________________________' if 'Signature' in label else ''
    sheet.merge_cells(f'B{row}:D{row}')

# FOOTER
row += 2
sheet.merge_cells(f'A{row}:D{row+1}')
footer = (
    "NOTES:\n"
    "1. This certificate issued under Clause 6.27 of JBCC Principal Building Agreement\n"
    "2. Copy of this certificate and snag list to be provided to Employer and Contractor\n"
    "3. Contractor has 10 working days to issue notice of dissatisfaction if they disagree\n"
    "4. All snag list items must be rectified during the Defects Liability Period"
)
sheet[f'A{row}'] = footer
sheet[f'A{row}'].font = Font(name='Arial', size=8, italic=True)
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 50

wb.save('/Users/donovan/Archive/AGENT/templates/JBCC_Practical_Completion_Certificate_Template.xlsx')
print("Practical Completion Certificate template created")
