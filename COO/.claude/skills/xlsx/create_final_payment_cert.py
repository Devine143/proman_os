from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = Workbook()
sheet = wb.active
sheet.title = "Final Payment Certificate"

sheet.column_dimensions['A'].width = 28
sheet.column_dimensions['B'].width = 35
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 18

header_font = Font(name='Arial', size=14, bold=True)
subheader_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
blue_font = Font(name='Arial', size=10, color='0000FF')
total_font = Font(name='Arial', size=11, bold=True)

header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
section_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
total_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# HEADER
sheet.merge_cells('A1:D1')
sheet['A1'] = 'FINAL PAYMENT CERTIFICATE'
sheet['A1'].font = Font(name='Arial', size=16, bold=True)
sheet['A1'].alignment = center_align
sheet['A1'].fill = header_fill

sheet.merge_cells('A2:D2')
sheet['A2'] = 'JBCC Principal Building Agreement'
sheet['A2'].font = subheader_font
sheet['A2'].alignment = center_align
sheet['A2'].fill = header_fill

sheet.merge_cells('A3:D3')
sheet['A3'] = 'Clause 6.28 - Final Account Payment'
sheet['A3'].font = Font(name='Arial', size=11, bold=True, color='FF0000')
sheet['A3'].alignment = center_align

# PROJECT INFO
row = 5
sheet[f'A{row}'] = 'PROJECT INFORMATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

for label, value in [
    ('Project Name:', '[Enter Project Name]'),
    ('Project Address:', '[Enter Project Address]'),
    ('Employer:', '[Enter Employer Name]'),
    ('Contractor:', '[Enter Contractor Name]'),
    ('Contract Date:', '[Enter Contract Date]'),
    ('Original Contract Sum (Excl VAT):', 0)
]:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'A{row}'].font = normal_font
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font
    if 'Sum' in label:
        sheet[f'B{row}'].number_format = 'R #,##0.00'
        sheet[f'B{row}'].fill = yellow_fill

# CERTIFICATE DETAILS
row += 2
sheet[f'A{row}'] = 'CERTIFICATE DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

for label, value, is_yellow in [
    ('Certificate Number:', 'FPC-001', True),
    ('Date of Final Completion Certificate:', '[Enter FC Cert Date]', True),
    ('Date of Final Payment Certificate:', datetime.now().strftime('%d %B %Y'), True),
    ('Payment Due Date:', '[21 days from certificate date]', True)
]:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'A{row}'].font = normal_font
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font if is_yellow else normal_font
    if is_yellow:
        sheet[f'B{row}'].fill = yellow_fill
    if 'Due Date' in label:
        sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True, color='FF0000')

# FINAL ACCOUNT SUMMARY
row += 2
sheet[f'A{row}'] = 'FINAL ACCOUNT SUMMARY'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
headers_row = row
for col, text in [('A', 'Description'), ('B', 'Contract Value'), ('C', 'Adjustments'), ('D', 'Final Value')]:
    sheet[f'{col}{row}'] = text
    sheet[f'{col}{row}'].font = subheader_font
    sheet[f'{col}{row}'].fill = header_fill
    sheet[f'{col}{row}'].alignment = center_align
    sheet[f'{col}{row}'].border = thin_border

row += 1
line1_row = row
items = [
    '1. Original Contract Works',
    '2. Approved Variations (Total)',
    '3. Provisional Sums Expended',
    '4. Prime Cost Sums Expended',
    '5. Dayworks',
    '6. Adjustments (Contract Price Fluctuations)'
]

for item in items:
    sheet[f'A{row}'] = item
    sheet[f'B{row}'] = 0
    sheet[f'C{row}'] = 0
    sheet[f'D{row}'] = f'=B{row}+C{row}'
    sheet[f'B{row}'].font = blue_font
    sheet[f'C{row}'].font = blue_font
    sheet[f'B{row}'].number_format = 'R #,##0.00'
    sheet[f'C{row}'].number_format = 'R #,##0.00'
    sheet[f'D{row}'].number_format = 'R #,##0.00'
    sheet[f'A{row}'].border = thin_border
    sheet[f'B{row}'].border = thin_border
    sheet[f'C{row}'].border = thin_border
    sheet[f'D{row}'].border = thin_border
    if '1. Original' in item:
        sheet[f'B{row}'] = '=B11'
        sheet[f'B{row}'].fill = yellow_fill
    row += 1

gross_val_row = row
sheet[f'A{row}'] = 'GROSS FINAL CONTRACT VALUE (Excl VAT)'
sheet[f'A{row}'].font = total_font
sheet[f'D{row}'] = f'=SUM(D{line1_row}:D{row-1})'
sheet[f'D{row}'].font = total_font
sheet[f'D{row}'].number_format = 'R #,##0.00'
sheet[f'A{row}'].fill = section_fill
sheet[f'D{row}'].fill = section_fill
sheet[f'A{row}'].border = thin_border
sheet[f'D{row}'].border = thin_border

# PAYMENTS & DEDUCTIONS
row += 2
sheet[f'A{row}'] = 'PAYMENTS TO DATE & DEDUCTIONS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
prev_certs_row = row
sheet[f'A{row}'] = 'Total Previous Interim Certificates (Excl VAT)'
sheet[f'D{row}'] = 0
sheet[f'D{row}'].font = blue_font
sheet[f'D{row}'].fill = yellow_fill
sheet[f'D{row}'].number_format = 'R #,##0.00'

row += 1
retention_held_row = row
sheet[f'A{row}'] = 'Retention Held to Date'
sheet[f'D{row}'] = 0
sheet[f'D{row}'].font = blue_font
sheet[f'D{row}'].fill = yellow_fill
sheet[f'D{row}'].number_format = 'R #,##0.00'

row += 1
deductions_row = row
sheet[f'A{row}'] = 'Other Deductions (if any)'
sheet[f'D{row}'] = 0
sheet[f'D{row}'].font = blue_font
sheet[f'D{row}'].number_format = 'R #,##0.00'

row += 2
retention_release_row = row
sheet[f'A{row}'] = 'Add: Final Retention Release'
sheet[f'D{row}'] = f'=D{retention_held_row}'
sheet[f'D{row}'].font = Font(name='Arial', size=10, color='008000')
sheet[f'D{row}'].number_format = 'R #,##0.00'

row += 2
net_excl_vat_row = row
sheet[f'A{row}'] = 'NET FINAL PAYMENT DUE (Excl VAT)'
sheet[f'A{row}'].font = total_font
sheet[f'D{row}'] = f'=D{gross_val_row}-D{prev_certs_row}-D{deductions_row}+D{retention_release_row}'
sheet[f'D{row}'].font = total_font
sheet[f'D{row}'].number_format = 'R #,##0.00'
sheet[f'A{row}'].fill = total_fill
sheet[f'D{row}'].fill = total_fill

# VAT
row += 2
vat_rate_row = row
sheet[f'A{row}'] = 'VAT Rate:'
sheet[f'C{row}'] = 0.15
sheet[f'C{row}'].font = blue_font
sheet[f'C{row}'].number_format = '0.0%'
sheet[f'C{row}'].fill = yellow_fill

row += 1
vat_row = row
sheet[f'A{row}'] = 'Add: VAT'
sheet[f'D{row}'] = f'=D{net_excl_vat_row}*$C${vat_rate_row}'
sheet[f'D{row}'].number_format = 'R #,##0.00'

row += 2
total_row = row
sheet[f'A{row}'] = 'TOTAL FINAL PAYMENT DUE (Incl VAT)'
sheet[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
sheet[f'D{row}'] = f'=D{net_excl_vat_row}+D{vat_row}'
sheet[f'D{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
sheet[f'D{row}'].number_format = 'R #,##0.00'
sheet[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sheet[f'D{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# LEGAL NOTICE
row += 3
sheet[f'A{row}'] = 'PAYMENT DUE DATE & LEGAL NOTICE'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row+2}')
legal_text = (
    "This FINAL PAYMENT CERTIFICATE concludes all payment obligations under the Contract. "
    "Payment is DUE WITHIN 21 DAYS from the date of this certificate.\n\n"
    "This certificate constitutes FINAL SETTLEMENT of all accounts between Employer and Contractor "
    "unless a notice of dissatisfaction is issued within 10 working days (Clause 6.28, JBCC). "
    "This certificate is BINDING AND EQUIVALENT TO CASH (Joob Joob Investments v Stocks Mavundla, 2009 SCA)."
)
sheet[f'A{row}'] = legal_text
sheet[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color='FF0000')
sheet[f'A{row}'].alignment = wrap_align
sheet[f'A{row}'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
sheet.row_dimensions[row].height = 60

# BANKING
row += 3
sheet[f'A{row}'] = 'CONTRACTOR BANKING DETAILS'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

for label in ['Bank Name:', 'Account Name:', 'Account Number:', 'Branch Code:']:
    row += 1
    sheet[f'A{row}'] = label
    sheet[f'B{row}'] = f'[Enter {label.replace(":", "")}]'
    sheet[f'B{row}'].font = blue_font

# CERTIFICATION
row += 3
sheet[f'A{row}'] = 'PRINCIPAL AGENT CERTIFICATION'
sheet[f'A{row}'].font = subheader_font
sheet[f'A{row}'].fill = section_fill
sheet.merge_cells(f'A{row}:D{row}')

row += 1
sheet.merge_cells(f'A{row}:D{row+2}')
cert_text = (
    "I hereby certify that:\n"
    "1. The Final Completion Certificate has been issued\n"
    "2. All defects have been rectified or arrangements made for their completion\n"
    "3. The final account has been prepared and agreed\n"
    "4. The amount stated represents the final settlement of all accounts under this Contract"
)
sheet[f'A{row}'] = cert_text
sheet[f'A{row}'].font = normal_font
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 60

row += 3
for label, value in [
    ('Principal Agent Name:', '[Enter PA Name]'),
    ('Company:', 'Proman Project Management'),
    ('Signature:', '________________________'),
    ('Date:', datetime.now().strftime('%d %B %Y'))
]:
    sheet[f'A{row}'] = label
    sheet[f'B{row}'] = value
    sheet[f'B{row}'].font = blue_font if '[' in str(value) or 'Proman' in str(value) or datetime.now().strftime('%Y') in str(value) else normal_font
    sheet.merge_cells(f'B{row}:D{row}')
    row += 1

# FOOTER
row += 1
sheet.merge_cells(f'A{row}:D{row+1}')
footer = (
    "NOTES:\n"
    "1. This certificate issued under Clause 6.28 (Final Payment Certificate) of JBCC PBA\n"
    "2. All retention has been released upon issuance of Final Completion Certificate\n"
    "3. This certificate constitutes final settlement - no further claims unless disputed within 10 days\n"
    "4. Contractor may claim loss of profit if payment not made within 21 days"
)
sheet[f'A{row}'] = footer
sheet[f'A{row}'].font = Font(name='Arial', size=8, italic=True)
sheet[f'A{row}'].alignment = wrap_align
sheet.row_dimensions[row].height = 55

wb.save('/Users/donovan/Archive/AGENT/templates/JBCC_Final_Payment_Certificate_Template.xlsx')
print("Final Payment Certificate template created")
