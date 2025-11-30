#!/usr/bin/env python3
"""
Combine Budget Analysis Excel Files into Single Workbook

This script combines multiple Excel budget analysis files into a single workbook
with separate sheets for each file, preserving all formulas and formatting.

Usage:
    python3 combine_budget_excel_files.py <project_name> <date_YYYYMMDD> <output_dir>

Example:
    python3 combine_budget_excel_files.py "11_Buck_Road" "20251119" "/path/to/output"
"""

import sys
from openpyxl import load_workbook, Workbook
import copy
from pathlib import Path

def combine_excel_files(project_name, date, output_dir):
    """
    Combine 4 budget analysis Excel files into one workbook.

    Args:
        project_name: Project name (e.g., "11_Buck_Road")
        date: Date in YYYYMMDD format (e.g., "20251119")
        output_dir: Directory containing the Excel files
    """

    output_path = Path(output_dir)

    # Define source files and their sheet names in combined workbook
    files = [
        (f"{project_name}_EVM_{date}.xlsx", "EVM Metrics"),
        (f"{project_name}_Variance_Analysis_{date}.xlsx", "Variance Analysis"),
        (f"{project_name}_Corrective_Actions_{date}.xlsx", "Corrective Actions"),
        (f"{project_name}_Cash_Flow_{date}.xlsx", "Cash Flow Forecast")
    ]

    # Create new workbook and remove default sheet
    wb_combined = Workbook()
    wb_combined.remove(wb_combined.active)

    print(f"Combining Excel files for {project_name} ({date})...")

    for filename, sheet_name in files:
        source_file = output_path / filename

        if not source_file.exists():
            print(f"  ⚠️  Warning: {filename} not found, skipping...")
            continue

        print(f"  ✓ Adding: {sheet_name}")

        # Load source workbook
        wb_source = load_workbook(str(source_file))
        ws_source = wb_source.active

        # Create new sheet in combined workbook
        ws_dest = wb_combined.create_sheet(sheet_name)

        # Copy all cell values and formulas
        for row in ws_source.iter_rows():
            for cell in row:
                ws_dest[cell.coordinate].value = cell.value

                # Copy cell formatting
                if cell.has_style:
                    ws_dest[cell.coordinate].font = copy.copy(cell.font)
                    ws_dest[cell.coordinate].border = copy.copy(cell.border)
                    ws_dest[cell.coordinate].fill = copy.copy(cell.fill)
                    ws_dest[cell.coordinate].number_format = copy.copy(cell.number_format)
                    ws_dest[cell.coordinate].protection = copy.copy(cell.protection)
                    ws_dest[cell.coordinate].alignment = copy.copy(cell.alignment)

        # Copy column widths
        for col in ws_source.column_dimensions:
            ws_dest.column_dimensions[col].width = ws_source.column_dimensions[col].width

        # Copy row heights
        for row in ws_source.row_dimensions:
            ws_dest.row_dimensions[row].height = ws_source.row_dimensions[row].height

        # Copy merged cells
        if ws_source.merged_cells:
            for merged_cell in ws_source.merged_cells.ranges:
                ws_dest.merge_cells(str(merged_cell))

    # Save combined workbook
    output_file = output_path / f"{project_name}_Budget_Analysis_Complete_{date}.xlsx"
    wb_combined.save(str(output_file))

    print(f"\n✅ Combined workbook created: {output_file.name}")
    print(f"   Sheets: {', '.join(wb_combined.sheetnames)}")
    print(f"   Size: {output_file.stat().st_size / 1024:.1f} KB")

    return str(output_file)

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    project_name = sys.argv[1]
    date = sys.argv[2]
    output_dir = sys.argv[3]

    combine_excel_files(project_name, date, output_dir)
